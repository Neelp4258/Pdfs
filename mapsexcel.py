import time
import re
import signal
import sys
import pandas as pd
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class GoogleMapsExcelExtractor:
    def __init__(self, headless=False):
        """Initialize the Google Maps extractor with Chrome driver"""
        self.options = webdriver.ChromeOptions()
        if headless:
            self.options.add_argument('--headless')
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument('--disable-blink-features=AutomationControlled')
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option('useAutomationExtension', False)
        # Suppress logging
        self.options.add_argument('--log-level=3')
        self.options.add_experimental_option('excludeSwitches', ['enable-logging'])
        
        self.driver = webdriver.Chrome(options=self.options)
        self.driver.maximize_window()  # Maximize for better element visibility
        self.wait = WebDriverWait(self.driver, 10)
        self.results = []
        self.stop_extraction = False  # Flag to control extraction
        self.search_query = ""  # Store the search query for filename
        
    def signal_handler(self, signum, frame):
        """Handle Ctrl+C signal gracefully"""
        print("\n\n🛑 STOPPING EXTRACTION...")
        print("Saving extracted data before closing...")
        self.stop_extraction = True
        
        # Immediately save data to prevent loss
        if self.results:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            filename = f"google_maps_results_stopped_{timestamp}.xlsx"
            success = self.save_to_excel_safe(filename)
            if success:
                print(f"✅ Data saved successfully to: {filename}")
                print(f"📊 Total extracted: {len(self.results)} results")
            else:
                print("❌ Failed to save data immediately")
        else:
            print("ℹ️ No data to save")
        
        # Close browser safely
        self.close_browser_safe()
        
        print("\n✅ Extraction stopped successfully!")
        print("Thank you for using Google Maps Excel Extractor!")
        sys.exit(0)
        
    def open_google_maps_and_wait(self):
        """Open Google Maps and wait for manual search"""
        try:
            # Navigate to Google Maps
            self.driver.get("https://www.google.com/maps")
            time.sleep(3)
            
            # Handle cookies/consent if present
            try:
                accept_buttons = self.driver.find_elements(By.XPATH, 
                    "//button[contains(text(), 'Accept') or contains(text(), 'Reject') or contains(text(), 'Got it')]")
                if accept_buttons:
                    accept_buttons[0].click()
                    time.sleep(1)
            except:
                pass
            
            print("\n" + "="*60)
            print("GOOGLE MAPS IS NOW OPEN")
            print("="*60)
            print("\nPlease manually:")
            print("1. Type your search query in the search box")
            print("2. Press ENTER to search")
            print("3. Wait for results to load")
            print("\nThen press ENTER here to start extracting data...")
            print("="*60)
            
            # Wait for user to press Enter and get search query
            search_input = input("\nEnter your search query for filename (optional) or press ENTER: ").strip()
            if search_input:
                self.search_query = search_input
            
            # Wait for user to confirm search is complete
            input("\nPress ENTER after you've completed your search and results are loaded...")
            
            # Verify that results are loaded
            try:
                self.wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div[role="feed"]'))
                )
                logger.info("Search results detected. Starting extraction...")
                return True
            except:
                logger.error("No search results found. Please make sure you've performed a search.")
                return False
                
        except Exception as e:
            logger.error(f"Error: {str(e)}")
            return False
    
    def extract_phone_from_text(self, text):
        """Extract phone numbers from text using regex"""
        if not text:
            return None
            
        # Clean the text
        text = text.strip()
        
        # Multiple phone patterns
        phone_patterns = [
            r'[\+]?[(]?[0-9]{1,3}[)]?[-\s\.]?[(]?[0-9]{1,4}[)]?[-\s\.]?[0-9]{1,4}[-\s\.]?[0-9]{1,9}',
            r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b',
            r'\b\d{10}\b'
        ]
        
        for pattern in phone_patterns:
            phones = re.findall(pattern, text)
            if phones:
                phone = phones[0].strip()
                if len(phone) >= 10:
                    return phone
        return None
    
    def extract_listing_details_from_panel(self):
        """Extract details from the currently open detail panel"""
        details = {
            'name': None,
            'phone': None,
            'email': None,
            'website': None,
            'address': None,
            'rating': None,
            'reviews_count': None,
            'category': None,
            'hours': None,
            'price_level': None
        }
        
        try:
            # Wait a bit for panel to fully load
            time.sleep(2)
            
            # Extract name - try multiple selectors
            name_selectors = [
                'h1.DUwDvf.fontHeadlineLarge',
                'h1[class*="fontHeadlineLarge"]',
                'h1.DUwDvf',
                '[role="main"] h1'
            ]
            
            for selector in name_selectors:
                try:
                    name_element = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if name_element and name_element.text:
                        details['name'] = name_element.text.strip()
                        break
                except:
                    continue
            
            # Extract category
            try:
                category_element = self.driver.find_element(By.CSS_SELECTOR, 
                    'button[jsaction*="category"] .DkEaL, .DkEaL')
                if category_element and category_element.text:
                    details['category'] = category_element.text.strip()
            except:
                pass
            
            # Extract all button/link texts that might contain info
            info_buttons = self.driver.find_elements(By.CSS_SELECTOR, 
                'button[data-item-id], button[data-tooltip], a[data-item-id]')
            
            for button in info_buttons:
                try:
                    item_id = button.get_attribute('data-item-id') or ''
                    aria_label = button.get_attribute('aria-label') or ''
                    text = button.text or ''
                    
                    # Phone extraction
                    if 'phone' in item_id.lower() or 'phone' in aria_label.lower():
                        if aria_label and ':' in aria_label:
                            phone_text = aria_label.split(':', 1)[1].strip()
                            details['phone'] = phone_text
                        elif text:
                            phone = self.extract_phone_from_text(text)
                            if phone:
                                details['phone'] = phone
                    
                    # Website extraction
                    elif 'website' in item_id.lower() or 'website' in aria_label.lower():
                        if text and ('.' in text or 'http' in text.lower()):
                            details['website'] = text.strip()
                    
                    # Address extraction
                    elif 'address' in item_id.lower() or 'address' in aria_label.lower():
                        if aria_label and ':' in aria_label:
                            details['address'] = aria_label.split(':', 1)[1].strip()
                        elif text:
                            details['address'] = text.strip()
                            
                except:
                    continue
            
            # Try alternative selectors for missing data
            if not details['phone']:
                try:
                    phone_links = self.driver.find_elements(By.CSS_SELECTOR, 'a[href^="tel:"]')
                    if phone_links:
                        phone_href = phone_links[0].get_attribute('href')
                        details['phone'] = phone_href.replace('tel:', '').strip()
                except:
                    pass
            
            # Extract rating and review count
            try:
                rating_element = self.driver.find_element(By.CSS_SELECTOR, 
                    'span[role="img"][aria-label*="stars"], span.MW4etd')
                rating_text = rating_element.get_attribute('aria-label') or rating_element.text
                if rating_text:
                    rating_match = re.search(r'([\d.]+)', rating_text)
                    if rating_match:
                        details['rating'] = rating_match.group(1)
                
                # Get review count
                review_element = self.driver.find_element(By.CSS_SELECTOR, 
                    'span.UY7F9 a span, .UY7F9')
                if review_element and review_element.text:
                    review_text = review_element.text
                    review_match = re.search(r'([\d,]+)', review_text)
                    if review_match:
                        details['reviews_count'] = review_match.group(1).replace(',', '')
            except:
                pass
            
            # Extract hours
            try:
                hours_element = self.driver.find_element(By.CSS_SELECTOR, 
                    '[data-item-id="oh"] .fontBodyMedium, .t39EBf .fontBodyMedium')
                if hours_element and hours_element.text:
                    details['hours'] = hours_element.text.strip()
            except:
                pass
            
            # Extract price level
            try:
                price_element = self.driver.find_element(By.CSS_SELECTOR, 
                    'span[aria-label*="Price"], .mgr77e .fontBodyMedium')
                if price_element:
                    price_text = price_element.get_attribute('aria-label') or price_element.text
                    if price_text and '$' in price_text:
                        details['price_level'] = price_text.strip()
            except:
                pass
            
            # Extract email from entire panel
            try:
                panel = self.driver.find_element(By.CSS_SELECTOR, 'div[role="main"]')
                panel_text = panel.text
                email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
                emails = re.findall(email_pattern, panel_text)
                if emails:
                    details['email'] = emails[0]
            except:
                pass
            
            logger.info(f"Extracted: {details.get('name', 'Unknown')} - Phone: {details.get('phone', 'N/A')}")
            
        except Exception as e:
            logger.error(f"Error extracting details: {str(e)}")
        
        return details
    
    def click_listing_by_index(self, index):
        """Click on a specific listing by index"""
        try:
            # Find the results panel
            results_panel = self.driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')
            
            # Find all listing links
            listings = results_panel.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
            
            if index >= len(listings):
                return False
            
            listing = listings[index]
            
            # Scroll into view
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", listing)
            time.sleep(1)
            
            # Click using JavaScript
            self.driver.execute_script("arguments[0].click();", listing)
            time.sleep(2)
            
            return True
            
        except Exception as e:
            logger.error(f"Error clicking listing {index}: {str(e)}")
            return False
    
    def get_total_results_count(self):
        """Get the total number of results currently loaded"""
        try:
            results_panel = self.driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')
            listings = results_panel.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
            return len(listings)
        except:
            return 0
    
    def scroll_results_panel(self):
        """Scroll the results panel to load more results"""
        try:
            results_panel = self.driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')
            
            # Get current count
            before_scroll = self.get_total_results_count()
            
            # Scroll to bottom
            self.driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", results_panel)
            time.sleep(3)
            
            # Check if new results loaded
            after_scroll = self.get_total_results_count()
            
            return after_scroll > before_scroll
            
        except Exception as e:
            logger.error(f"Error scrolling: {str(e)}")
            return False
    
    def extract_all_results(self, max_results=None):
        """Extract contact information from all search results"""
        processed_indices = set()
        consecutive_failures = 0
        no_new_results_count = 0
        
        print("\n" + "="*60)
        print("EXTRACTION IN PROGRESS")
        print("="*60)
        print("Press CTRL+C at any time to STOP and SAVE extracted data")
        print("The script will continue extracting ALL results until stopped")
        print("="*60 + "\n")
        
        try:
            while not self.stop_extraction:  # Check stop flag
                total_listings = self.get_total_results_count()
                
                if total_listings == 0:
                    logger.error("No listings found!")
                    break
                
                # Process each listing
                new_results_found = False
                for i in range(total_listings):
                    # Check stop flag before processing each listing
                    if self.stop_extraction:
                        logger.info("Stopping extraction as requested...")
                        break
                        
                    if i in processed_indices:
                        continue
                    
                    logger.info(f"Processing listing {i+1}/{total_listings} (Total extracted: {len(self.results)})")
                    
                    try:
                        # Click on the listing
                        if self.click_listing_by_index(i):
                            # Extract details
                            details = self.extract_listing_details_from_panel()
                            
                            if details['name']:
                                self.results.append(details)
                                processed_indices.add(i)
                                consecutive_failures = 0
                                new_results_found = True
                            else:
                                consecutive_failures += 1
                                logger.warning(f"No data extracted for listing {i+1}")
                        else:
                            consecutive_failures += 1
                    except Exception as e:
                        logger.error(f"Error processing listing {i+1}: {str(e)}")
                        consecutive_failures += 1
                        # If browser connection is lost, stop extraction
                        if "connection" in str(e).lower() or "session" in str(e).lower():
                            logger.warning("Browser connection lost, stopping extraction...")
                            break
                    
                    # If too many failures, try scrolling
                    if consecutive_failures > 3:
                        logger.info("Multiple failures, trying to scroll for more results...")
                        try:
                            if not self.scroll_results_panel():
                                no_new_results_count += 1
                            else:
                                no_new_results_count = 0
                            consecutive_failures = 0
                        except Exception as e:
                            logger.error(f"Error scrolling: {str(e)}")
                            if "connection" in str(e).lower() or "session" in str(e).lower():
                                logger.warning("Browser connection lost during scrolling, stopping extraction...")
                                break
                
                # Check stop flag before continuing
                if self.stop_extraction:
                    break
                
                # Always try to load more results
                logger.info(f"Scrolling to load more results... (Total extracted so far: {len(self.results)})")
                try:
                    if not self.scroll_results_panel():
                        no_new_results_count += 1
                        logger.info("No new results loaded this time")
                    else:
                        no_new_results_count = 0
                        logger.info("New results loaded, continuing extraction...")
                except Exception as e:
                    logger.error(f"Error during final scroll: {str(e)}")
                    if "connection" in str(e).lower() or "session" in str(e).lower():
                        logger.warning("Browser connection lost, stopping extraction...")
                        break
                
                # If we've tried multiple times and no new results, we're done
                if no_new_results_count > 3:
                    logger.info(f"No more results available. Extraction complete!")
                    logger.info(f"Total results extracted: {len(self.results)}")
                    break
                
                # Small delay to prevent overwhelming the server
                time.sleep(1)
                
        except Exception as e:
            logger.error(f"Extraction error: {str(e)}")
            if "connection" in str(e).lower() or "session" in str(e).lower():
                logger.warning("Browser connection lost during extraction")
        
        return self.results
    
    def save_to_excel_safe(self, filename='google_maps_results.xlsx'):
        """Save extracted results to an Excel file with advanced formatting"""
        if not self.results:
            logger.warning("No results to save")
            return False
        
        try:
            # Convert to DataFrame
            df = pd.DataFrame(self.results)
            
            # Clean and process data
            df = self.clean_dataframe(df)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Google Maps Data', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Google Maps Data']
                
                # Apply comprehensive formatting
                self.apply_excel_formatting(workbook, worksheet, df)
                
                # Create summary sheet
                self.create_summary_sheet(workbook, df)
            
            print(f"✅ Results saved to {filename}")
            return True
            
        except Exception as e:
            print(f"❌ Error saving to Excel: {str(e)}")
            return False
    
    def clean_dataframe(self, df):
        """Clean and process the DataFrame"""
        # Clean phone numbers
        if 'phone' in df.columns:
            df['phone'] = df['phone'].astype(str).apply(self.clean_phone_number)
        
        # Clean ratings
        if 'rating' in df.columns:
            df['rating'] = pd.to_numeric(df['rating'], errors='coerce')
        
        # Clean review counts
        if 'reviews_count' in df.columns:
            df['reviews_count'] = pd.to_numeric(df['reviews_count'], errors='coerce')
        
        # Fill NaN values
        df = df.fillna('')
        
        # Add extraction metadata
        df['extraction_date'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
        if self.search_query:
            df['search_query'] = self.search_query
        
        return df
    
    def clean_phone_number(self, phone):
        """Clean and format phone numbers"""
        if pd.isna(phone) or phone == '' or phone == 'nan':
            return ''
        
        # Remove non-digit characters except +
        cleaned = re.sub(r'[^\d+]', '', str(phone))
        
        # Format if it's a valid length
        if len(cleaned) >= 10:
            return cleaned
        
        return str(phone)
    
    def apply_excel_formatting(self, workbook, worksheet, df):
        """Apply comprehensive Excel formatting"""
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        column_widths = {
            'name': 25,
            'phone': 15,
            'email': 25,
            'website': 30,
            'address': 35,
            'rating': 10,
            'reviews_count': 12,
            'category': 20,
            'hours': 20,
            'price_level': 12,
            'extraction_date': 18,
            'search_query': 20
        }
        
        for idx, column in enumerate(df.columns, 1):
            column_letter = worksheet.cell(row=1, column=idx).column_letter
            width = column_widths.get(column, 15)
            worksheet.column_dimensions[column_letter].width = width
        
        # Apply borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border
        
        # Alternate row colors
        light_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        for row_num in range(2, len(df) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_num, column=col_num).fill = light_fill
        
        # Format rating column with color scale
        if 'rating' in df.columns:
            rating_col_idx = list(df.columns).index('rating') + 1
            rating_col_letter = worksheet.cell(row=1, column=rating_col_idx).column_letter
            rating_range = f"{rating_col_letter}2:{rating_col_letter}{len(df)+1}"
            
            # Add color scale rule
            color_scale_rule = ColorScaleRule(
                start_type='min', start_color='FF6B6B',
                mid_type='percentile', mid_value=50, mid_color='FFE66D',
                end_type='max', end_color='4ECDC4'
            )
            worksheet.conditional_formatting.add(rating_range, color_scale_rule)
        
        # Add data bars for review count
        if 'reviews_count' in df.columns and df['reviews_count'].notna().any():
            reviews_col_idx = list(df.columns).index('reviews_count') + 1
            reviews_col_letter = worksheet.cell(row=1, column=reviews_col_idx).column_letter
            reviews_range = f"{reviews_col_letter}2:{reviews_col_letter}{len(df)+1}"
            
            data_bar_rule = DataBarRule(
                start_type='min', end_type='max',
                color='5B9BD5', showValue=True
            )
            worksheet.conditional_formatting.add(reviews_range, data_bar_rule)
        
        # Freeze panes
        worksheet.freeze_panes = 'A2'
        
        # Add table formatting
        if len(df) > 0:
            table_range = f"A1:{worksheet.cell(row=len(df)+1, column=len(df.columns)).coordinate}"
            table = Table(displayName="GoogleMapsData", ref=table_range)
            
            # Add table style
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            worksheet.add_table(table)
    
    def create_summary_sheet(self, workbook, df):
        """Create a summary statistics sheet"""
        summary_sheet = workbook.create_sheet("Summary Statistics")
        
        # Summary statistics
        stats = [
            ["Summary Statistics", ""],
            ["", ""],
            ["Total Records", len(df)],
            ["Records with Phone", df['phone'].astype(str).str.len().gt(0).sum()],
            ["Records with Email", df['email'].astype(str).str.len().gt(0).sum()],
            ["Records with Website", df['website'].astype(str).str.len().gt(0).sum()],
            ["Records with Rating", df['rating'].notna().sum()],
            ["", ""],
            ["Rating Statistics", ""],
            ["Average Rating", df['rating'].mean() if df['rating'].notna().any() else 0],
            ["Highest Rating", df['rating'].max() if df['rating'].notna().any() else 0],
            ["Lowest Rating", df['rating'].min() if df['rating'].notna().any() else 0],
            ["", ""],
            ["Extraction Info", ""],
            ["Extraction Date", pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Search Query", self.search_query if self.search_query else "Not provided"]
        ]
        
        # Write statistics
        for row_idx, (label, value) in enumerate(stats, 1):
            summary_sheet.cell(row=row_idx, column=1, value=label)
            summary_sheet.cell(row=row_idx, column=2, value=value)
        
        # Format summary sheet
        summary_sheet.column_dimensions['A'].width = 20
        summary_sheet.column_dimensions['B'].width = 25
        
        # Header formatting
        header_cell = summary_sheet.cell(row=1, column=1)
        header_cell.font = Font(bold=True, size=14, color='1F4E79')
        
        # Make category headers bold
        category_rows = [1, 9, 14]
        for row in category_rows:
            cell = summary_sheet.cell(row=row, column=1)
            cell.font = Font(bold=True, color='1F4E79')
    
    def save_to_excel(self, filename='google_maps_results.xlsx'):
        """Save extracted results to an Excel file"""
        return self.save_to_excel_safe(filename)
    
    def close_browser_safe(self):
        """Safely close the browser without throwing errors"""
        try:
            if hasattr(self, 'driver') and self.driver:
                self.driver.quit()
                print("🔒 Browser closed successfully")
        except Exception as e:
            print(f"ℹ️ Browser was already closed or connection lost")
    
    def close(self):
        """Close the browser"""
        self.close_browser_safe()

# Example usage
if __name__ == "__main__":
    # Initialize the extractor
    extractor = GoogleMapsExcelExtractor(headless=False)
    
    # Set up signal handler for graceful shutdown
    signal.signal(signal.SIGINT, extractor.signal_handler)
    
    try:
        # Open Google Maps and wait for manual search
        if extractor.open_google_maps_and_wait():
            # Extract results
            print("\nStarting extraction process...")
            print("This will extract ALL available results until completion or manual stop (CTRL+C)")
            
            # Extract all results
            results = extractor.extract_all_results()
            
            # Generate filename based on search query and timestamp
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            if extractor.search_query:
                # Clean search query for filename
                clean_query = re.sub(r'[^\w\s-]', '', extractor.search_query)
                clean_query = re.sub(r'[-\s]+', '_', clean_query)
                filename = f"google_maps_{clean_query}_{timestamp}.xlsx"
            else:
                filename = f"google_maps_results_{timestamp}.xlsx"
            
            if extractor.save_to_excel(filename):
                # Print summary
                print(f"\n{'='*60}")
                print(f"EXTRACTION COMPLETE ✅")
                print(f"{'='*60}")
                print(f"Extracted {len(results)} total results")
                print(f"Results saved to: {filename}")
                print(f"📊 Excel file includes:")
                print(f"   • Main data sheet with professional formatting")
                print(f"   • Summary statistics sheet")
                print(f"   • Color-coded ratings and data bars")
                print(f"   • Filterable table format")
                
                if results:
                    print(f"\nFirst 5 results preview:")
                    for i, result in enumerate(results[:5]):
                        print(f"\n{i+1}. {result['name']}")
                        if result['phone']:
                            print(f"   📞 Phone: {result['phone']}")
                        if result['email']:
                            print(f"   📧 Email: {result['email']}")
                        if result['website']:
                            print(f"   🌐 Website: {result['website']}")
                        if result['rating']:
                            print(f"   ⭐ Rating: {result['rating']} stars")
                        if result['category']:
                            print(f"   🏷️ Category: {result['category']}")
                        if result['address']:
                            print(f"   📍 Address: {result['address']}")
            else:
                print("❌ Failed to save results to Excel")
        else:
            print("Failed to detect search results. Please try again.")
        
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        # Try to save whatever we have
        if extractor.results:
            timestamp = time.strftime("%Y%m%d_%H%M%S") 
            filename = f"google_maps_results_emergency_{timestamp}.xlsx"
            if extractor.save_to_excel_safe(filename):
                print(f"Emergency save completed: {filename}")
            else:
                print("❌ Emergency save failed")
        
    finally:
        # Only proceed if we haven't already handled Ctrl+C
        if not extractor.stop_extraction:
            # Always save data if we have any
            if extractor.results:
                print(f"\n📊 Total extracted: {len(extractor.results)} results")
            
            print(f"\n{'='*60}")
            print("CLEANUP")
            print(f"{'='*60}")
            
            input("Press ENTER to close the browser...")
            extractor.close()
            print("✅ Browser closed successfully!")
            print("Thank you for using Google Maps Excel Extractor!")